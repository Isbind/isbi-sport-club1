from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

def creer_fichier_excel():
    # Créer un nouveau classeur
    wb = Workbook()
    
    # Supprimer la feuille par défaut si elle existe
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Fonction pour formater les en-têtes
    def formater_en_tete(worksheet):
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Ajuster la largeur des colonnes
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width
    
    # 1. ONGLET "MEMBRES"
    ws_membres = wb.create_sheet("Membres")
    en_tetes_membres = ["ID", "Nom", "Prénom", "Date de naissance", "Téléphone", 
                        "Email", "Date d'inscription", "Statut"]
    ws_membres.append(en_tetes_membres)
    
    # 2. ONGLET "ABONNEMENTS"
    ws_abonnements = wb.create_sheet("Abonnements")
    en_tetes_abonnements = ["ID Membre", "Type d'abonnement", "Date début", 
                           "Date fin", "Prix", "Statut"]
    ws_abonnements.append(en_tetes_abonnements)
    
    # 3. ONGLET "PAIEMENTS"
    ws_paiements = wb.create_sheet("Paiements")
    en_tetes_paiements = ["N° Paiement", "ID Membre", "Date", "Montant", 
                         "Méthode", "Statut", "Notes"]
    ws_paiements.append(en_tetes_paiements)
    
    # 4. ONGLET "PRÉSENCES"
    ws_presences = wb.create_sheet("Présences")
    en_tetes_presences = ["Date", "ID Membre", "Heure d'arrivée", 
                         "Heure de départ", "Durée", "Coach"]
    ws_presences.append(en_tetes_presences)
    
    # Appliquer le formatage à tous les onglets
    for ws in wb.worksheets:
        formater_en_tete(ws)
    
    # Sauvegarder le fichier avec la date actuelle
    date_actuelle = datetime.now().strftime("%Y%m%d")
    nom_fichier = f"ISBISPORTCLUB_Suivi_{date_actuelle}.xlsx"
    wb.save(nom_fichier)
    
    print(f"Fichier Excel '{nom_fichier}' créé avec succès !")

if __name__ == "__main__":
    creer_fichier_excel()
