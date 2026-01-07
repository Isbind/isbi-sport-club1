import os
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, Series, PieChart
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image
from datetime import datetime, timedelta
import random

# Configuration des styles
HEADER_FILL = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
TITLE_FONT = Font(name='Arial', size=14, bold=True)
NORMAL_FONT = Font(name='Arial', size=10)
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def creer_fichier_excel():
    # Création du classeur Excel
    wb = Workbook()
    
    # Suppression de la feuille par défaut
    for sheet in wb.sheetnames:
        wb.remove(wb[sheet])
    
    # Création des onglets
    membres = wb.create_sheet("Membres")
    abonnements = wb.create_sheet("Abonnements")
    paiements = wb.create_sheet("Paiements")
    presences = wb.create_sheet("Présences")
    tableau_bord = wb.create_sheet("Tableau de Bord")
    
    # Configuration des en-têtes
    configurer_onglet_membres(membres)
    configurer_onglet_abonnements(abonnements)
    configurer_onglet_paiements(paiements)
    configurer_onglet_presences(presences)
    configurer_tableau_bord(tableau_bord, wb)
    
    # Ajustement automatique des largeurs de colonnes
    for sheet in wb.sheetnames:
        ajuster_largeur_colonnes(wb[sheet])
    
    # Protection des feuilles
    for sheet in wb.sheetnames:
        if sheet != "Tableau de Bord":  # On ne protège pas le tableau de bord
            wb[sheet].protection.sheet = True
            wb[sheet].protection.formatCells = False
    
    # Sauvegarde du fichier
    nom_fichier = f"ISBISPORTCLUB_Suivi_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(nom_fichier)
    print(f"Fichier Excel créé avec succès : {nom_fichier}")

def configurer_onglet_membres(ws):
    # Titre
    ws['A1'] = "GESTION DES MEMBRES"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:H1')
    
    # En-têtes
    en_tetes = ["ID", "Nom", "Prénom", "Téléphone", "Email", "Date d'inscription", "Date de naissance", "Médecine du sport"]
    for col, en_tete in enumerate(en_tetes, 1):
        cell = ws.cell(row=3, column=col, value=en_tete)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
    
    # Données d'exemple
    for i in range(1, 21):
        ws.cell(row=i+3, column=1, value=f"M{i:03d}")
        ws.cell(row=i+3, column=2, value=f"Nom{i}")
        ws.cell(row=i+3, column=3, value=f"Prénom{i}")
        ws.cell(row=i+3, column=4, value=f"06{random.randint(10000000, 99999999)}")
        ws.cell(row=i+3, column=5, value=f"membre{i}@email.com")
        ws.cell(row=i+3, column=6, value=datetime.now() - timedelta(days=random.randint(1, 365)))
        ws.cell(row=i+3, column=7, value=datetime.now() - timedelta(days=random.randint(18*365, 65*365)))
        ws.cell(row=i+3, column=8, value=random.choice(["Oui", "Non"]))
    
    # Format des dates
    for row in ws.iter_rows(min_row=4, min_col=6, max_col=7, max_row=23):
        for cell in row:
            cell.number_format = 'DD/MM/YYYY'
    
    # Ajout de la validation des données pour la médecine du sport
    dv = DataValidation(type="list", formula1='"Oui,Non"', showDropDown=True)
    ws.add_data_validation(dv)
    dv.add(f'H4:H{ws.max_row}')

def configurer_onglet_abonnements(ws):
    # Titre
    ws['A1'] = "GESTION DES ABONNEMENTS"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:H1')
    
    # En-têtes
    en_tetes = ["ID Abonnement", "ID Membre", "Type", "Date début", "Date fin", "Prix", "Statut", "Méthode de paiement"]
    for col, en_tete in enumerate(en_tetes, 1):
        cell = ws.cell(row=3, column=col, value=en_tete)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
    
    # Données d'exemple
    for i in range(1, 21):
        date_debut = datetime.now() - timedelta(days=random.randint(1, 365))
        type_abo = random.choice(["Mensuel", "Trimestriel", "Annuel"])
        duree = 30 if type_abo == "Mensuel" else 90 if type_abo == "Trimestriel" else 365
        prix = 30 if type_abo == "Mensuel" else 80 if type_abo == "Trimestriel" else 300
        
        ws.cell(row=i+3, column=1, value=f"A{i:04d}")
        ws.cell(row=i+3, column=2, value=f"M{random.randint(1, 20):03d}")
        ws.cell(row=i+3, column=3, value=type_abo)
        ws.cell(row=i+3, column=4, value=date_debut)
        ws.cell(row=i+3, column=5, value=date_debut + timedelta(days=duree))
        ws.cell(row=i+3, column=6, value=prix)
        ws.cell(row=i+3, column=7, value=random.choice(["Actif", "Expiré", "Résilié"]))
        ws.cell(row=i+3, column=8, value=random.choice(["Espèces", "Carte", "Chèque", "Virement"]))
    
    # Format des dates et des nombres
    for row in ws.iter_rows(min_row=4, min_col=4, max_col=5, max_row=23):
        for cell in row:
            cell.number_format = 'DD/MM/YYYY'
    
    for row in ws.iter_rows(min_row=4, min_col=6, max_col=6, max_row=23):
        for cell in row:
            cell.number_format = '#,##0.00 €'
    
    # Validation des données
    dv_type = DataValidation(type="list", formula1='"Mensuel,Trimestriel,Annuel"', showDropDown=True)
    dv_statut = DataValidation(type="list", formula1='"Actif,Expiré,Résilié"', showDropDown=True)
    dv_paiement = DataValidation(type="list", formula1='"Espèces,Carte,Chèque,Virement"', showDropDown=True)
    
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_statut)
    ws.add_data_validation(dv_paiement)
    
    dv_type.add(f'C4:C{ws.max_row}')
    dv_statut.add(f'G4:G{ws.max_row}')
    dv_paiement.add(f'H4:H{ws.max_row}')

def configurer_onglet_paiements(ws):
    # Titre
    ws['A1'] = "SUIVI DES PAIEMENTS"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')
    
    # En-têtes
    en_tetes = ["ID Paiement", "ID Membre", "Date", "Montant", "Méthode", "Statut"]
    for col, en_tete in enumerate(en_tetes, 1):
        cell = ws.cell(row=3, column=col, value=en_tete)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
    
    # Données d'exemple
    for i in range(1, 51):
        date_paiement = datetime.now() - timedelta(days=random.randint(1, 365))
        montant = random.choice([30, 80, 300, 35, 85, 310])
        
        ws.cell(row=i+3, column=1, value=f"P{i:05d}")
        ws.cell(row=i+3, column=2, value=f"M{random.randint(1, 20):03d}")
        ws.cell(row=i+3, column=3, value=date_paiement)
        ws.cell(row=i+3, column=4, value=montant)
        ws.cell(row=i+3, column=5, value=random.choice(["Espèces", "Carte", "Chèque"]))
        ws.cell(row=i+3, column=6, value="Payé")
    
    # Format des dates et des nombres
    for row in ws.iter_rows(min_row=4, min_col=3, max_col=3, max_row=53):
        for cell in row:
            cell.number_format = 'DD/MM/YYYY'
    
    for row in ws.iter_rows(min_row=4, min_col=4, max_col=4, max_row=53):
        for cell in row:
            cell.number_format = '#,##0.00 €'
    
    # Validation des données
    dv_methode = DataValidation(type="list", formula1='"Espèces,Carte,Chèque,Virement"', showDropDown=True)
    dv_statut = DataValidation(type="list", formula1='"Payé,En attente,Annulé"', showDropDown=True)
    
    ws.add_data_validation(dv_methode)
    ws.add_data_validation(dv_statut)
    
    dv_methode.add(f'E4:E{ws.max_row}')
    dv_statut.add(f'F4:F{ws.max_row}')

def configurer_onglet_presences(ws):
    # Titre
    ws['A1'] = "SUIVI DES PRÉSENCES"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:G1')
    
    # En-têtes
    en_tetes = ["ID Séance", "ID Membre", "Date", "Heure d'arrivée", "Heure de départ", "Durée (min)", "Activité"]
    for col, en_tete in enumerate(en_tetes, 1):
        cell = ws.cell(row=3, column=col, value=en_tete)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
    
    # Données d'exemple
    for i in range(1, 101):
        date_seance = datetime.now() - timedelta(days=random.randint(1, 30))
        heure_arrivee = datetime.combine(date_seance.date(), datetime.min.time()) + timedelta(hours=random.randint(8, 20), minutes=random.randint(0, 59))
        duree = random.randint(30, 120)
        heure_depart = heure_arrivee + timedelta(minutes=duree)
        
        ws.cell(row=i+3, column=1, value=f"S{i:05d}")
        ws.cell(row=i+3, column=2, value=f"M{random.randint(1, 20):03d}")
        ws.cell(row=i+3, column=3, value=date_seance)
        ws.cell(row=i+3, column=4, value=heure_arrivee)
        ws.cell(row=i+3, column=5, value=heure_depart)
        ws.cell(row=i+3, column=6, value=duree)
        ws.cell(row=i+3, column=7, value=random.choice(["Musculation", "Cours collectif", "Cardio", "Étirement"]))
    
    # Format des dates et heures
    for row in ws.iter_rows(min_row=4, min_col=3, max_col=3, max_row=103):
        for cell in row:
            cell.number_format = 'DD/MM/YYYY'
    
    for row in ws.iter_rows(min_row=4, min_col=4, max_col=5, max_row=103):
        for cell in row:
            cell.number_format = 'HH:MM'
    
    # Formule pour calculer la durée
    for row in range(4, 104):
        ws[f'F{row}'] = f'=HOUR(E{row}-D{row})*60 + MINUTE(E{row}-D{row})'
    
    # Validation des données pour l'activité
    dv_activite = DataValidation(type="list", formula1='"Musculation,Cours collectif,Cardio,Étirement"', showDropDown=True)
    ws.add_data_validation(dv_activite)
    dv_activite.add(f'G4:G{ws.max_row}')

def configurer_tableau_bord(ws, wb):
    # Titre
    ws['A1'] = "TABLEAU DE BORD ISBISPORTCLUB"
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color='1F4E78')
    ws.merge_cells('A1:H1')
    
    # Sous-titre
    ws['A2'] = f"Données au {datetime.now().strftime('%d/%m/%Y')}"
    ws['A2'].font = Font(name='Arial', size=12, italic=True, color='7F7F7F')
    ws.merge_cells('A2:H2')
    
    # Statistiques clés
    stats = [
        ["Membres actifs", '=COUNTIF(Membres!H:H,"Oui")'],
        ["Abonnements actifs", '=COUNTIF(Abonnements!G:G,"Actif")'],
        ["Recettes mensuelles", '=SUMIFS(Paiements!D:D,Paiements!C:C,">="&EOMONTH(TODAY(),-1)+1,Paiements!C:C,"<="&EOMONTH(TODAY(),0))'],
        ["Taux de fréquentation", '=COUNT(Présences!A:A)/MAX(1,COUNT(Membres!A:A))/30']
    ]
    
    for i, stat in enumerate(stats, 1):
        ws[f'A{4 + i}'] = stat[0]
        ws[f'B{4 + i}'] = stat[1]
        ws[f'B{4 + i}'].number_format = '#,##0.00' if i == 4 else '#,##0'
    
    # Graphique des abonnements par type
    ws['A10'] = "Répartition des abonnements par type"
    ws['A10'].font = Font(name='Arial', size=12, bold=True)
    
    # Création d'un tableau croisé dynamique (simulé)
    ws['A12'] = "Type d'abonnement"
    ws['B12'] = "Nombre"
    ws['A13'] = "Mensuel"
    ws['A14'] = "Trimestriel"
    ws['A15'] = "Annuel"
    
    ws['B13'] = '=COUNTIF(Abonnements!C:C,"Mensuel")'
    ws['B14'] = '=COUNTIF(Abonnements!C:C,"Trimestriel")'
    ws['B15'] = '=COUNTIF(Abonnements!C:C,"Annuel")'
    
    # Création du graphique
    pie = PieChart()
    labels = Reference(ws, min_col=1, min_row=13, max_row=15)
    data = Reference(ws, min_col=2, min_row=12, max_row=15)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Répartition des abonnements"
    ws.add_chart(pie, "D12")
    
    # Graphique des recettes mensuelles
    ws['A25'] = "Recettes mensuelles (12 derniers mois)"
    ws['A25'].font = Font(name='Arial', size=12, bold=True)
    
    # Tableau des données (simulé)
    ws['A27'] = "Mois"
    ws['B27'] = "Montant"
    
    for i in range(12):
        mois = (datetime.now().month - 1 - i) % 12 + 1
        annee = datetime.now().year if (datetime.now().month - i) > 0 else datetime.now().year - 1
        ws[f'A{28 + i}'] = f"{mois:02d}/{annee}"
        ws[f'B{28 + i}'] = f'=SOMMEPROD((MOIS(Paiements!C$4:C$1000)={mois})*(ANNEE(Paiements!C$4:C$1000)={annee})*Paiements!D$4:D$1000)'
    
    # Création du graphique à barres
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Recettes mensuelles"
    bar.x_axis.title = "Mois"
    bar.y_axis.title = "Montant (€)"
    
    data = Reference(ws, min_col=2, min_row=27, max_row=39)
    cats = Reference(ws, min_col=1, min_row=28, max_row=39)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    
    ws.add_chart(bar, "D25")
    
    # Mise en forme
    for row in ws[3:39]:
        for cell in row:
            cell.border = BORDER
            if cell.row == 3 or cell.column == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

def ajuster_largeur_colonnes(ws):
    # Désactiver temporairement l'ajustement automatique des colonnes
    # pour éviter les problèmes avec les cellules fusionnées
    for i, col in enumerate(ws.columns, 1):
        # Largeur fixe pour toutes les colonnes
        ws.column_dimensions[get_column_letter(i)].width = 15

if __name__ == "__main__":
    creer_fichier_excel()
